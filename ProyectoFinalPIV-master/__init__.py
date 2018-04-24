from flask import Flask,request,render_template,jsonify,url_for
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime

import openpyxl



app=Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI']='sqlite:///Diccionario.db'
db=SQLAlchemy(app)





class Categoria(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    nombre=db.Column(db.String(50),nullable=False)

class Diccionario(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    palabra=db.Column(db.String(50),nullable=False)
    significado=db.Column(db.String(100),nullable=False)
    id_categoria=db.Column(db.Integer, db.ForeignKey('categoria.id'), nullable=False)
    categoria=db.relationship('Categoria',backref=db.backref('palabras', lazy=True))
    fecha_creacion=db.Column(db.DateTime, default=datetime.now())



@app.route('/',methods=['GET','POST'])
def index():
    dato = {'ninguno': 'Ningún Resultado Encontrado...', 'th': '', 'object': []}
    if request.method == 'POST' and len(request.form["word_to_search"])>0:
        variable=request.form['word_to_search']
        dato['th'] = ['Id', 'Palabra', 'Significado', 'Fecha de Ingreso']
        dato['td'] = ['id', 'palabra', 'significado', 'fecha_creacion']
        dato['object'] = Diccionario.query.filter(Diccionario.palabra.like("%" + variable + "%")).all()
        dato['ninguno'] = ''
        dato['cantidad_resultados'] = len(dato['object'])
        dato['palabra_buscada'] = variable
    if len(dato['object']) == 0:
        #dato.clear()
        dato['ninguno'] = 'Ningún Resultado Encontrado...'
    return render_template('index.html',dato=dato)



@app.route('/API/<string:word>',methods=['GET','POST'])
@app.route('/API')
def api(word='casa'):
    datos=Diccionario.query.filter(Diccionario.palabra.like("%" + word + "%")).all()
    i=0
    diccionario={}
    for elementos in datos:
        diccionario[i]={'palabra':elementos.palabra,'traduccion':elementos.significado}
        i+=1
    return  jsonify({'resultados':diccionario})

@app.route('/CargarDiccionario',methods=['GET','POST'])
def CargarDiccionario():
    doc=openpyxl.load_workbook('es_gu.xlsx')
    hoja=doc.get_sheet_by_name('Hoja1')
    i=1
    for numero in range(1747):
        palabra=hoja.cell(row=i,column=1).value
        significado=hoja.cell(row=i,column=2).value
        i+=1
        if palabra !=None and significado !=None:
            dato=Diccionario(palabra=palabra,significado=significado,id_categoria=1)
            db.session.add(dato)
            db.session.commit()
            print('Ingresando: ', palabra)
    return 'Diccionario Cargado'


if __name__=='__main__':

    app.run(debug=True, port=1234)